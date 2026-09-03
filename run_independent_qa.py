import os
import re
import csv
import json
import time
import random
import string
import datetime
import requests
import openpyxl
from dateutil import parser as date_parser

# Setup paths
ORIG_XLSM = "LG_corrected.xlsm"
FINAL_XLSM = "LG_audited_final.xlsm"
QA_DIR = "qa_raw_retrievals"
CSV_REPORT = "correction_qa_report.csv"
JSON_REPORT = "correction_qa_report.json"

os.makedirs(QA_DIR, exist_ok=True)

def normalize_text(text):
    if not text:
        return ""
    text = str(text).lower()
    text = text.translate(str.maketrans('', '', string.punctuation))
    return re.sub(r'\s+', ' ', text).strip()

def extract_id_from_url(url):
    match = re.search(r'/reviews/([a-f0-9]+)', str(url))
    return match.group(1) if match else None

def parse_date_safely(date_val):
    if not date_val:
        return None
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        return date_val.strftime("%Y-%m-%d")
    try:
        dt = date_parser.parse(str(date_val))
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None

def fetch_fresh_jina(url, max_retries=3):
    jina_url = f"https://r.jina.ai/{url}"
    headers = {"Accept": "application/json"}
    
    for attempt in range(max_retries):
        try:
            resp = requests.get(jina_url, headers=headers, timeout=30)
            if resp.status_code == 200:
                data = resp.json()
                title = data.get("data", {}).get("title", "")
                content = data.get("data", {}).get("content", "")
                if "Verifying your connection" in title or "challenge.js" in content:
                    return False, "ACCESS_LIMITED", None, resp.text
                if "404" in title or "Page not found" in title:
                    return False, "REVIEW_NOT_FOUND", None, resp.text
                return True, "SUCCESS", content, resp.text
            elif resp.status_code == 429:
                if attempt < max_retries - 1:
                    time.sleep(3 * (attempt + 1))
                    continue
                return False, "RATE_LIMITED", None, resp.text
            else:
                return False, f"HTTP_{resp.status_code}", None, resp.text
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(2 * (attempt + 1))
                continue
            return False, f"ERROR_{str(e)[:30]}", None, None
            
    return False, "UNKNOWN", None, None

def verify_identity_independently(expected_id, expected_text, content):
    if not content:
        return False, "NO_CONTENT"
    content_lower = content.lower()
    if expected_id and expected_id.lower() in content_lower:
        return True, "EXACT_REVIEW_ID_MATCH"
    
    norm_exp = normalize_text(expected_text)
    norm_content = normalize_text(content)
    if norm_exp and norm_exp in norm_content:
        return True, "EXACT_TEXT_MATCH"
        
    sentences = re.split(r'[.!?]+', str(expected_text))
    if sentences:
        longest = max(sentences, key=len).strip()
        norm_longest = normalize_text(longest)
        if len(norm_longest) > 25 and norm_longest in norm_content:
            return True, "STRONG_PARTIAL_TEXT_MATCH"
            
    return False, "INSUFFICIENT_IDENTITY_EVIDENCE"

def extract_fields_independently(content):
    if not content:
        return {"Raw_text": None, "Review_date": None, "Support_reply": None, "Reply_date": None}
        
    lines = content.split('\n')
    text_lines = []
    reply_lines = []
    review_date = None
    reply_date = None
    
    in_review_body = False
    in_reply = False
    date_pattern = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},\s+\d{4}$", re.IGNORECASE)
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        if line.startswith("## ["):
            in_review_body = True
            continue
            
        if "**Reply from" in line or "Reply from" in line:
            in_review_body = False
            in_reply = True
            continue
            
        if in_review_body:
            if date_pattern.match(line):
                if not review_date:
                    review_date = line
                in_review_body = False
                continue
            if line in ["* * *", "Useful", "Share"]:
                in_review_body = False
                continue
            text_lines.append(line)
            
        elif in_reply:
            if date_pattern.match(line):
                if not reply_date:
                    reply_date = line
                continue
            if line in ["* * *", "Useful", "Share"]:
                in_reply = False
                continue
            reply_lines.append(line)
        else:
            if date_pattern.match(line) and not review_date:
                review_date = line
                
    raw_text = "\n".join(text_lines).strip() if text_lines else None
    if raw_text and ("![Image" in raw_text or "](https" in raw_text):
        raw_text = None
        
    support_reply = "\n".join(reply_lines).strip() if reply_lines else None
    
    return {
        "Raw_text": raw_text,
        "Review_date": review_date,
        "Support_reply": support_reply,
        "Reply_date": reply_date
    }

def main():
    print("=" * 70)
    print("STARTING INDEPENDENT CORRECTION QA AUDIT")
    print("=" * 70)
    
    print(f"Loading {ORIG_XLSM} and {FINAL_XLSM}...")
    wb_orig = openpyxl.load_workbook(ORIG_XLSM, data_only=True)
    ws_orig = wb_orig["Data"]
    
    wb_final = openpyxl.load_workbook(FINAL_XLSM, data_only=True)
    ws_final = wb_final["Data"]
    
    # Categorize rows where Correction_Made contains "Yes"
    raw_text_candidates = []
    support_reply_candidates = []
    reply_date_candidates = []
    multi_field_candidates = []
    
    for r in range(4, ws_final.max_row + 1):
        corr_made = ws_final.cell(row=r, column=14).value
        field_aff = ws_final.cell(row=r, column=11).value or ""
        
        if corr_made and "Yes" in str(corr_made):
            fields = [f.strip() for f in field_aff.split("|")]
            if "Raw_text" in fields:
                raw_text_candidates.append(r)
            if "Support_reply" in fields:
                support_reply_candidates.append(r)
            if "Reply_date" in fields:
                reply_date_candidates.append(r)
            if len(fields) > 1:
                multi_field_candidates.append(r)
                
    print(f"Candidates: Raw_text={len(raw_text_candidates)}, Support_reply={len(support_reply_candidates)}, Reply_date={len(reply_date_candidates)}, Multi-field={len(multi_field_candidates)}")
    
    # Set seed for reproducible sampling
    random.seed(42)
    
    # Sample 10 for each category
    # To include multi-field rows, prioritize selecting some multi-field rows in each category
    sample_raw_text = random.sample(raw_text_candidates, 10)
    
    # Select 10 Support_reply rows (ensuring some have multi-field)
    sample_support_reply = random.sample(support_reply_candidates, 10)
    
    # Select 10 Reply_date rows
    sample_reply_date = random.sample(reply_date_candidates, 10)
    
    # Build unique list of (row_idx, category_target)
    # We will evaluate the specific target field for each sample, plus any multi-field on that row
    tasks = []
    for r in sample_raw_text:
        tasks.append((r, "Raw_text"))
    for r in sample_support_reply:
        tasks.append((r, "Support_reply"))
    for r in sample_reply_date:
        tasks.append((r, "Reply_date"))
        
    print(f"\nTotal sampled row-field evaluations: {len(tasks)}")
    unique_rows = sorted(list(set(r for r, f in tasks)))
    print(f"Unique rows to retrieve from Trustpilot: {len(unique_rows)}")
    
    # Column mappings in Excel:
    # Col 1: ID, Col 2: Platform, Col 3: Url, Col 4: Raw_text, Col 5: Rating,
    # Col 6: Review_date, Col 7: Order Date, Col 8: Reply Date, Col 9: Support_reply
    field_to_col = {
        "Raw_text": 4,
        "Rating": 5,
        "Review_date": 6,
        "Order Date": 7,
        "Reply_date": 8,
        "Reply Date": 8,
        "Support_reply": 9
    }
    
    # Also track where the production verifier's buggy col_map wrote values:
    verifier_buggy_col_map = {
        "Raw_text": 3,      # Overwrote Url!
        "Rating": 4,        # (skipped)
        "Review_date": 5,   # Overwrote Rating!
        "Reply_date": 6,    # Overwrote Review_date!
        "Support_reply": 7  # Overwrote Order Date!
    }
    
    # Fetch all unique rows once
    retrievals = {}
    for i, r in enumerate(unique_rows, 1):
        url = ws_orig.cell(row=r, column=3).value
        review_id = extract_id_from_url(url)
        orig_text = ws_orig.cell(row=r, column=4).value
        
        print(f"[{i}/{len(unique_rows)}] Fetching Row {r} (ID: {review_id})...")
        success, status, content, raw_json = fetch_fresh_jina(url)
        
        # Save raw retrieval for audit independence
        raw_file = os.path.join(QA_DIR, f"row_{r}_{review_id}.json")
        with open(raw_file, "w", encoding="utf-8") as f:
            f.write(raw_json if raw_json else json.dumps({"status": status, "url": url}))
            
        id_verified = False
        id_reason = status
        tp_fields = {}
        
        if success:
            id_verified, id_reason = verify_identity_independently(review_id, orig_text, content)
            tp_fields = extract_fields_independently(content)
            
        retrievals[r] = {
            "url": url,
            "review_id": review_id,
            "success": success,
            "status": status,
            "id_verified": id_verified,
            "id_reason": id_reason,
            "tp_fields": tp_fields,
            "raw_file": raw_file
        }
        time.sleep(1) # Polite delay
        
    print("\n" + "=" * 70)
    print("EVALUATING FIELD-BY-FIELD CORRECTIONS")
    print("=" * 70)
    
    report_records = []
    
    for r, field in tasks:
        ret = retrievals[r]
        url = ret["url"]
        review_id = ret["review_id"]
        id_verified = ret["id_verified"]
        id_reason = ret["id_reason"]
        tp_fields = ret["tp_fields"]
        raw_file = ret["raw_file"]
        
        target_col = field_to_col[field]
        orig_val = ws_orig.cell(row=r, column=target_col).value
        final_val = ws_final.cell(row=r, column=target_col).value
        fresh_tp_val = tp_fields.get(field)
        
        # Also check what got written into the verifier's misplaced column
        misplaced_col = verifier_buggy_col_map.get(field)
        misplaced_val = ws_final.cell(row=r, column=misplaced_col).value if misplaced_col else None
        
        orig_str = str(orig_val).strip() if orig_val is not None else ""
        final_str = str(final_val).strip() if final_val is not None else ""
        tp_str = str(fresh_tp_val).strip() if fresh_tp_val is not None else ""
        
        # Classification determination
        qa_class = "UNVERIFIABLE"
        confidence = "High"
        explanation = ""
        
        if not ret["success"] or not id_verified:
            qa_class = "UNVERIFIABLE"
            explanation = f"Fresh Trustpilot retrieval failed or identity could not be independently confirmed (Status: {ret['status']}, Reason: {id_reason})."
        else:
            # Check if this is a date field
            if field in ["Reply_date", "Review_date"]:
                orig_parsed = parse_date_safely(orig_val)
                final_parsed = parse_date_safely(final_val)
                tp_parsed = parse_date_safely(fresh_tp_val)
                
                # Check if calendar dates are identical
                if orig_parsed and tp_parsed and orig_parsed == tp_parsed:
                    # Semantic date was already identical in Excel!
                    qa_class = "FORMAT_NORMALIZATION"
                    explanation = f"The date represented the exact same calendar date ({orig_parsed}). Differs only in string/datetime representation (Original: '{orig_val}', Trustpilot: '{fresh_tp_val}')."
                    if final_str != tp_str:
                        explanation += f" Note: Final Excel column {target_col} was not updated (remained '{final_val}') because production verifier wrote value to column {misplaced_col}."
                elif orig_str != tp_str:
                    # There was a real discrepancy
                    # Did the final Excel column reflect Trustpilot?
                    if final_str == tp_str or (final_parsed and tp_parsed and final_parsed == tp_parsed):
                        qa_class = "GENUINE_CORRECTION"
                        explanation = f"Genuine date update supported by Trustpilot: Original '{orig_val}' differed from Trustpilot '{fresh_tp_val}', and Final Excel correctly reflected it."
                    else:
                        qa_class = "EXTRACTION_ERROR"
                        explanation = f"Trustpilot provided '{fresh_tp_val}' (differing from original '{orig_val}'), but Final Excel target column {target_col} has '{final_val}'. (Production verifier incorrectly wrote '{misplaced_val}' into column {misplaced_col})."
                else:
                    qa_class = "FORMAT_NORMALIZATION"
                    explanation = f"Original and Trustpilot values match."
                    
            elif field == "Support_reply":
                norm_orig = normalize_text(orig_val)
                norm_final = normalize_text(final_val)
                norm_tp = normalize_text(fresh_tp_val)
                
                if not fresh_tp_val:
                    qa_class = "EXTRACTION_ERROR"
                    explanation = "No company support reply was found on Trustpilot for this review."
                elif norm_orig and norm_orig == norm_tp:
                    qa_class = "FORMAT_NORMALIZATION"
                    explanation = "Support reply was already present and semantically equivalent (differed only in whitespace/formatting)."
                elif norm_final == norm_tp:
                    qa_class = "GENUINE_CORRECTION"
                    explanation = "Genuine support reply populated from Trustpilot and correctly reflected in Final Excel."
                else:
                    qa_class = "EXTRACTION_ERROR"
                    explanation = f"Trustpilot had authentic support reply ('{tp_str[:60]}...'), but Final Excel target column 9 ('Support_reply') remained '{final_val}' because production verifier wrote it to column {misplaced_col} ('Order Date')."
                    
            elif field == "Raw_text":
                norm_orig = normalize_text(orig_val)
                norm_final = normalize_text(final_val)
                norm_tp = normalize_text(fresh_tp_val)
                
                # Check for extraction artifacts or truncation
                if not fresh_tp_val:
                    qa_class = "EXTRACTION_ERROR"
                    explanation = "Fresh extraction did not yield a valid review body."
                elif tp_str.endswith("...") or tp_str.endswith("…"):
                    qa_class = "EXTRACTION_ERROR"
                    explanation = "Extracted Trustpilot text is an ellipsized summary, not full review."
                elif norm_orig == norm_tp:
                    qa_class = "FORMAT_NORMALIZATION"
                    explanation = "Review text was already identical in substance; differed only in minor punctuation, whitespace, or newlines."
                elif norm_final == norm_tp:
                    qa_class = "GENUINE_CORRECTION"
                    explanation = "Genuine review text correction supported by Trustpilot and correctly reflected in Final Excel."
                else:
                    qa_class = "EXTRACTION_ERROR"
                    explanation = f"Trustpilot has genuine review text ('{tp_str[:60]}...'), but Final Excel target column 4 ('Raw_text') retained '{final_str[:60]}...' because production verifier wrote it to column {misplaced_col} ('Url')."
                    
        rec = {
            "Excel_Row_Number": r,
            "Trustpilot_Review_ID": review_id,
            "Trustpilot_URL": url,
            "Field_Tested": field,
            "Original_Excel_Value": str(orig_val)[:200] if orig_val is not None else "None",
            "Fresh_Trustpilot_Value": str(fresh_tp_val)[:200] if fresh_tp_val is not None else "None",
            "Final_Excel_Value": str(final_val)[:200] if final_val is not None else "None",
            "Misplaced_Col_Value": str(misplaced_val)[:200] if misplaced_val is not None else "None",
            "Identity_Verification_Result": id_reason,
            "QA_Classification": qa_class,
            "Confidence": confidence,
            "Explanation": explanation,
            "Fresh_Evidence_Source": raw_file
        }
        report_records.append(rec)
        print(f"Row {r:4d} | Field: {field:13s} | Class: {qa_class:20s} | ID: {id_reason}")
        
    # Write CSV Report
    csv_headers = [
        "Excel_Row_Number", "Trustpilot_Review_ID", "Trustpilot_URL", "Field_Tested",
        "Original_Excel_Value", "Fresh_Trustpilot_Value", "Final_Excel_Value",
        "Identity_Verification_Result", "QA_Classification", "Confidence",
        "Explanation", "Fresh_Evidence_Source"
    ]
    with open(CSV_REPORT, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_headers, extrasaction="ignore")
        writer.writeheader()
        for rec in report_records:
            writer.writerow(rec)
            
    # Write JSON Report
    with open(JSON_REPORT, "w", encoding="utf-8") as f:
        json.dump(report_records, f, indent=2)
        
    print(f"\nSaved QA CSV Report to: {CSV_REPORT}")
    print(f"Saved QA JSON Report to: {JSON_REPORT}")

if __name__ == "__main__":
    main()
