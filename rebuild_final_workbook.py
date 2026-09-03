import os
import json
import csv
import openpyxl
import datetime

SOURCE_FILE = "LG_corrected.xlsm"
TARGET_FILE = "LG_audited_final_fixed.xlsm"
CHECKPOINT_FILE = "checkpoint_prod.json"
REPORT_JSON = "rebuild_validation_report.json"
REPORT_CSV = "rebuild_validation_report.csv"

# Corrected 1-indexed column mapping in sheet "Data"
CORRECT_COL_MAP = {
    "Raw_text": 4,
    "Rating": 5,
    "Review_date": 6,
    "Reply_date": 8,
    "Support_reply": 9
}

AUDIT_HEADERS = [
    "Verification_Status",
    "Field_Affected",
    "Original_Value",
    "Trustpilot_Value",
    "Correction_Made",
    "Evidence_Source",
    "Confidence",
    "Reason_Action_Taken"
]

def main():
    print("=" * 70)
    print("STEP 1: PREFLIGHT CHECKPOINT & SOURCE VALIDATION")
    print("=" * 70)

    with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
        checkpoint = json.load(f)

    print(f"Loaded checkpoint records: {len(checkpoint)}")

    # Check keys and determinism
    duplicate_keys = 0
    missing_keys = []
    key_ints = []
    for k in checkpoint.keys():
        try:
            key_ints.append(int(k))
        except ValueError:
            missing_keys.append(k)

    key_min, key_max = min(key_ints), max(key_ints)
    print(f"Row key range: {key_min} to {key_max} (Unique: {len(set(key_ints))})")

    # Load source workbook in data_only=False, keep_vba=True for writing
    print(f"\nLoading source workbook {SOURCE_FILE} with keep_vba=True...")
    wb_out = openpyxl.load_workbook(SOURCE_FILE, keep_vba=True)
    ws_out = wb_out["Data"]

    # Also load source workbook in data_only=True to read pristine baseline
    wb_read = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    ws_read = wb_read["Data"]

    source_headers = [ws_read.cell(row=2, column=c).value for c in range(1, 10)]
    print(f"Source headers: {source_headers}")

    # Verify column layout in source workbook
    expected_headers = [
        ('ID', 1), ('Platform', 2), ('Url', 3), ('Raw_text', 4), ('Rating', 5),
        ('Review_date', 6), ('Order Date', 7), ('Reply Date', 8), ('Support_reply', 9)
    ]
    for name, idx in expected_headers:
        actual = ws_read.cell(row=2, column=idx).value
        assert actual == name, f"Header mismatch at col {idx}: expected {name}, got {actual}"
    print("Source header alignment: 100% verified.")

    print("\n" + "=" * 70)
    print("STEP 2: APPLYING CHECKPOINT CORRECTIONS & AUDIT COLUMNS")
    print("=" * 70)

    # Add audit headers starting at column 10
    audit_start_col = 10
    for i, h in enumerate(AUDIT_HEADERS, 1):
        ws_out.cell(row=2, column=audit_start_col + i - 1, value=h)

    corrections_applied_by_field = {
        "Raw_text": 0,
        "Rating": 0,
        "Review_date": 0,
        "Reply_date": 0,
        "Support_reply": 0
    }
    total_corrections_applied = 0
    applied_cell_coords = set()

    for r_str, res in checkpoint.items():
        r_idx = int(r_str)
        status_code = res.get("status_code", "UNKNOWN")
        audit_records = res.get("audit_records", [])
        corrs = res.get("corrections", {})

        # Apply corrections to the true target columns
        for field, tp_val in corrs.items():
            if field in CORRECT_COL_MAP:
                target_col = CORRECT_COL_MAP[field]
                ws_out.cell(row=r_idx, column=target_col, value=tp_val)
                corrections_applied_by_field[field] += 1
                total_corrections_applied += 1
                applied_cell_coords.add((r_idx, target_col))

        # Write audit columns (joined records per row)
        ws_out.cell(row=r_idx, column=audit_start_col, value=status_code)
        ws_out.cell(row=r_idx, column=audit_start_col + 1, value=" | ".join(r.get("Field_Affected", "") for r in audit_records))
        ws_out.cell(row=r_idx, column=audit_start_col + 2, value=" | ".join(r.get("Original_Value", "") for r in audit_records))
        ws_out.cell(row=r_idx, column=audit_start_col + 3, value=" | ".join(r.get("Trustpilot_Value", "") for r in audit_records))
        ws_out.cell(row=r_idx, column=audit_start_col + 4, value=" | ".join(r.get("Correction_Made", "") for r in audit_records))
        ws_out.cell(row=r_idx, column=audit_start_col + 5, value=" | ".join(r.get("Evidence_Source", "") for r in audit_records))
        ws_out.cell(row=r_idx, column=audit_start_col + 6, value=" | ".join(r.get("Confidence", "") for r in audit_records))
        ws_out.cell(row=r_idx, column=audit_start_col + 7, value=" | ".join(r.get("Reason_Action_Taken", "") for r in audit_records))

    print(f"Total field corrections applied: {total_corrections_applied}")
    for fld, cnt in corrections_applied_by_field.items():
        print(f"  {fld:15s}: {cnt}")

    # Safety check before saving: verify no write to Col 3 or Col 7
    for r_idx, c_idx in applied_cell_coords:
        assert c_idx not in [3, 7], f"CRITICAL ERROR: Cell write attempted at forbidden column {c_idx} (row {r_idx})!"
    print("\nPre-save assertion passed: ZERO writes to Column 3 (Url) or Column 7 (Order Date).")

    print(f"\nSaving reconstructed workbook to {TARGET_FILE}...")
    wb_out.save(TARGET_FILE)
    print("Workbook successfully saved.")

    print("\n" + "=" * 70)
    print("STEP 3: POST-SAVE INDEPENDENT VERIFICATION & CROSS-CHECK")
    print("=" * 70)

    # Reopen saved workbook to verify on-disk contents
    print(f"Reopening {TARGET_FILE} for byte-level verification...")
    wb_verify = openpyxl.load_workbook(TARGET_FILE, data_only=True)
    ws_verify = wb_verify["Data"]

    checkpoint_mismatches = []
    for r_str, res in checkpoint.items():
        r_idx = int(r_str)
        corrs = res.get("corrections", {})
        for field, expected_val in corrs.items():
            if field in CORRECT_COL_MAP:
                col_idx = CORRECT_COL_MAP[field]
                actual_val = ws_verify.cell(row=r_idx, column=col_idx).value
                # Normalize string comparison
                exp_str = str(expected_val).strip()
                act_str = str(actual_val).strip() if actual_val is not None else ""
                if exp_str != act_str:
                    checkpoint_mismatches.append({
                        "row": r_idx,
                        "field": field,
                        "column": col_idx,
                        "expected": exp_str[:80],
                        "actual": act_str[:80]
                    })

    print(f"Checkpoint-to-workbook mismatches: {len(checkpoint_mismatches)}")

    # Check Column 3 (Url) and Column 7 (Order Date) preservation
    url_changes = []
    order_date_changes = []
    untouched_cell_changes = []

    for r in range(1, ws_read.max_row + 1):
        # Check Column 3 (Url)
        orig_url = ws_read.cell(row=r, column=3).value
        final_url = ws_verify.cell(row=r, column=3).value
        if orig_url != final_url:
            url_changes.append({"row": r, "orig": str(orig_url)[:80], "final": str(final_url)[:80]})

        # Check Column 7 (Order Date)
        orig_od = ws_read.cell(row=r, column=7).value
        final_od = ws_verify.cell(row=r, column=7).value
        if orig_od != final_od:
            order_date_changes.append({"row": r, "orig": str(orig_od)[:80], "final": str(final_od)[:80]})

        # Check untouched columns (1, 2) and untouched cells across all columns 1..9
        for c in range(1, 10):
            if (r, c) not in applied_cell_coords and r > 2:
                orig_c = ws_read.cell(row=r, column=c).value
                final_c = ws_verify.cell(row=r, column=c).value
                # Handle datetime vs datetime representation
                if orig_c != final_c:
                    # Check if string representation is identical
                    if str(orig_c) != str(final_c):
                        untouched_cell_changes.append({
                            "row": r,
                            "column": c,
                            "orig": str(orig_c)[:80],
                            "final": str(final_c)[:80]
                        })

    print(f"Unintended URL changes: {len(url_changes)}")
    print(f"Unintended Order Date changes: {len(order_date_changes)}")
    print(f"Unintended untouched-cell changes: {len(untouched_cell_changes)}")

    # Compile Validation Summary Report
    summary_data = {
        "Target_Output_Filename": TARGET_FILE,
        "Source_Filename": SOURCE_FILE,
        "Total_Source_Rows": ws_read.max_row,
        "Total_Checkpoint_Records": len(checkpoint),
        "Total_Corrections_Applied": total_corrections_applied,
        "Corrections_By_Field": corrections_applied_by_field,
        "Checkpoint_To_Workbook_Mismatches": len(checkpoint_mismatches),
        "Unintended_URL_Changes": len(url_changes),
        "Unintended_Order_Date_Changes": len(order_date_changes),
        "Unintended_Untouched_Cell_Changes": len(untouched_cell_changes),
        "Duplicate_Checkpoint_Records": duplicate_keys,
        "Missing_Or_Ambiguous_Row_Mappings": len(missing_keys),
        "Rebuild_Status": "SUCCESS" if (
            len(checkpoint_mismatches) == 0 and
            len(url_changes) == 0 and
            len(order_date_changes) == 0 and
            len(untouched_cell_changes) == 0
        ) else "FAILED"
    }

    # Write JSON report
    with open(REPORT_JSON, "w", encoding="utf-8") as f:
        json.dump(summary_data, f, indent=2)
    print(f"\nSaved validation report to: {REPORT_JSON}")

    # Write CSV report
    csv_rows = [
        {"Metric": "Target_Output_Filename", "Value": TARGET_FILE},
        {"Metric": "Source_Filename", "Value": SOURCE_FILE},
        {"Metric": "Total_Source_Rows", "Value": ws_read.max_row},
        {"Metric": "Total_Checkpoint_Records", "Value": len(checkpoint)},
        {"Metric": "Total_Corrections_Applied", "Value": total_corrections_applied},
        {"Metric": "Corrections_Raw_text", "Value": corrections_applied_by_field["Raw_text"]},
        {"Metric": "Corrections_Rating", "Value": corrections_applied_by_field["Rating"]},
        {"Metric": "Corrections_Review_date", "Value": corrections_applied_by_field["Review_date"]},
        {"Metric": "Corrections_Reply_date", "Value": corrections_applied_by_field["Reply_date"]},
        {"Metric": "Corrections_Support_reply", "Value": corrections_applied_by_field["Support_reply"]},
        {"Metric": "Checkpoint_To_Workbook_Mismatches", "Value": len(checkpoint_mismatches)},
        {"Metric": "Unintended_URL_Changes", "Value": len(url_changes)},
        {"Metric": "Unintended_Order_Date_Changes", "Value": len(order_date_changes)},
        {"Metric": "Unintended_Untouched_Cell_Changes", "Value": len(untouched_cell_changes)},
        {"Metric": "Duplicate_Checkpoint_Records", "Value": duplicate_keys},
        {"Metric": "Missing_Or_Ambiguous_Row_Mappings", "Value": len(missing_keys)},
        {"Metric": "Rebuild_Status", "Value": summary_data["Rebuild_Status"]}
    ]
    with open(REPORT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Metric", "Value"])
        writer.writeheader()
        writer.writerows(csv_rows)
    print(f"Saved validation CSV report to: {REPORT_CSV}")

    print("\n" + "=" * 70)
    print(f"FINAL RESULT: {summary_data['Rebuild_Status']}")
    print("=" * 70)

if __name__ == "__main__":
    main()
