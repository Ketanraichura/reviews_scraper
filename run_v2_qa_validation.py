import random
import openpyxl

def main():
    random.seed(42)
    
    wb_orig = openpyxl.load_workbook("LG_corrected.xlsm", data_only=True)
    ws_orig = wb_orig["Data"]
    
    wb_v2 = openpyxl.load_workbook("LG_audited_v2.xlsm", data_only=True)
    ws_v2 = wb_v2["Data"]
    
    rec_rating = []
    rec_order_date = []
    sup_reply_disc = []
    reply_date_disc = []
    raw_text_disc = []
    review_date_disc = []
    clean = []
    unverified = []
    rate_limited = []
    
    for r in range(4, ws_v2.max_row + 1):
        status = ws_v2.cell(row=r, column=10).value
        verdicts = ws_v2.cell(row=r, column=11).value or ""
        recovered = ws_v2.cell(row=r, column=13).value or ""
        disc = ws_v2.cell(row=r, column=12).value or ""
        made_corr = ws_v2.cell(row=r, column=14).value
        
        if status == "IDENTITY_UNVERIFIED":
            unverified.append(r)
        elif status == "RATE_LIMITED_OR_BLOCKED":
            rate_limited.append(r)
        elif status == "VERIFIED":
            if "Rating" in recovered:
                rec_rating.append(r)
            if "Order Date" in recovered:
                rec_order_date.append(r)
            if "Support_reply" in disc or "Support_reply" in recovered:
                sup_reply_disc.append(r)
            if "Reply Date" in disc or "Reply Date" in recovered:
                reply_date_disc.append(r)
            if "Raw_text" in disc:
                raw_text_disc.append(r)
            if "Review_date" in disc:
                review_date_disc.append(r)
                
            if made_corr == "No" and not "DISCREPANCY" in verdicts and not "RECOVERED" in verdicts:
                clean.append(r)
                
    def sample_it(lst, n):
        return random.sample(lst, min(n, len(lst)))
        
    samples = {
        "Recovered Rating": sample_it(rec_rating, 20),
        "Recovered Order Date": sample_it(rec_order_date, 20),
        "Support Reply Changes": sample_it(sup_reply_disc, 20),
        "Reply Date Changes": sample_it(reply_date_disc, 20),
        "Raw Text Discrepancies": sample_it(raw_text_disc, 20),
        "Review Date True Discrepancies": sample_it(review_date_disc, 20),
        "Clean Perfect Matches": sample_it(clean, 10),
        "Identity Unverified": sample_it(unverified, 10),
        "Rate Limited": sample_it(rate_limited, 10),
    }
    
    print("\nSample Sizes:")
    for k, v in samples.items():
        print(f"  {k}: {len(v)}")
        
    with open("qa_audit_samples.txt", "w") as out:
        out.write("V2 QA AUDIT SAMPLES\n")
        out.write("="*70 + "\n\n")
        
        for category, r_idxs in samples.items():
            out.write(f"\n{category.upper()}\n")
            out.write("-" * 50 + "\n")
            for r_int in r_idxs:
                out.write(f"Row {r_int}:\n")
                if "Rating" in category:
                    out.write(f"  Orig: {ws_orig.cell(row=r_int, column=5).value} -> V2: {ws_v2.cell(row=r_int, column=5).value}\n")
                elif "Order Date" in category:
                    out.write(f"  Orig: {ws_orig.cell(row=r_int, column=7).value} -> V2: {ws_v2.cell(row=r_int, column=7).value}\n")
                elif "Support Reply" in category:
                    orig_val = ws_orig.cell(row=r_int, column=9).value
                    v2_val = ws_v2.cell(row=r_int, column=9).value
                    out.write(f"  Orig: {repr(str(orig_val)[:80])}... -> V2: {repr(str(v2_val)[:80])}...\n")
                elif "Reply Date" in category:
                    out.write(f"  Orig: {ws_orig.cell(row=r_int, column=8).value} -> V2: {ws_v2.cell(row=r_int, column=8).value}\n")
                elif "Raw Text" in category:
                    orig_val = ws_orig.cell(row=r_int, column=4).value
                    v2_val = ws_v2.cell(row=r_int, column=4).value
                    out.write(f"  Orig: {repr(str(orig_val)[:80])}... -> V2: {repr(str(v2_val)[:80])}...\n")
                elif "Review Date" in category:
                    out.write(f"  Orig: {ws_orig.cell(row=r_int, column=6).value} -> V2: {ws_v2.cell(row=r_int, column=6).value}\n")
                else:
                    out.write(f"  Status: {ws_v2.cell(row=r_int, column=10).value}\n")
            out.write("\n")
            
    print("Exported samples to qa_audit_samples.txt")

if __name__ == "__main__":
    main()
