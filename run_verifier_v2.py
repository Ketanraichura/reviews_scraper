import os
import csv
import json
import concurrent.futures
from collections import Counter
import openpyxl

from src_v2.extractor import extract_all_fields
from src_v2.comparator import compare_all
from src_v2.fetcher import fetch_jina

SOURCE_FILE = "LG_corrected.xlsm"
TARGET_FILE = "LG_audited_v2.xlsm"

def extract_id_from_url(url):
    import re
    match = re.search(r'/reviews/([a-f0-9]+)', str(url))
    return match.group(1) if match else None

def process_row(row_idx, row_data):
    url = row_data.get("Url")
    expected_id = extract_id_from_url(url)
    expected_text = row_data.get("Raw_text")
    
    if not url:
        return row_idx, {"status_code": "NO_URL", "results": {}}
        
    jina_res, from_cache = fetch_jina(url, expected_id, max_retries=3)
    status = jina_res["status"]
    data = jina_res.get("data")
    
    if status != "SUCCESS" or not data:
        return row_idx, {"status_code": status, "results": {}}
        
    content = data.get("content", "")
    title_meta = data.get("title", "")
    
    ext_fields = extract_all_fields(content, expected_id, expected_text, title_meta)
    
    if not ext_fields["Identity_Confirmed"]:
        return row_idx, {"status_code": "IDENTITY_UNVERIFIED", "reason": ext_fields.get("Reason"), "results": {}}
        
    # Compare
    results = compare_all(row_data, ext_fields)
    
    return row_idx, {"status_code": "VERIFIED", "results": results}

def main():
    print("=" * 70)
    print("VERIFIER V2 ENGINE")
    print("=" * 70)
    
    wb = openpyxl.load_workbook(SOURCE_FILE, keep_vba=True)
    ws = wb["Data"]
    
    wb_read = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    ws_read = wb_read["Data"]
    
    headers = [ws_read.cell(row=2, column=c).value for c in range(1, 10)]
    print(f"Headers: {headers}")
    
    tasks = []
    for r in range(4, ws_read.max_row + 1):
        if ws_read.cell(row=r, column=2).value == "Trustpilot":
            row_data = {headers[c-1]: ws_read.cell(row=r, column=c).value for c in range(1, 10)}
            tasks.append((r, row_data))
            
    print(f"Found {len(tasks)} Trustpilot rows to process.")
    
    processed_results = {}
    
    # Process concurrently (mostly cached, so 5 workers is fine; fetch_jina handles 429s)
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        futures = {executor.submit(process_row, r_idx, r_data): r_idx for r_idx, r_data in tasks}
        for i, future in enumerate(concurrent.futures.as_completed(futures), 1):
            r_idx, res = future.result()
            processed_results[r_idx] = res
            if i % 100 == 0:
                print(f"Processed {i} / {len(tasks)} rows...")
                
    # Now write to workbook
    print("\nWriting Audit Output...")
    
    AUDIT_HEADERS = [
        "Verification_Status",
        "Field_Verdicts",
        "Real_Discrepancies_Found",
        "Fields_Recovered",
        "Correction_Made",
        "Normalization_Applied",
        "Audit_Details"
    ]
    
    start_col = 10
    for i, h in enumerate(AUDIT_HEADERS):
        ws.cell(row=2, column=start_col + i, value=h)
        
    col_map = {
        "Raw_text": 4,
        "Rating": 5,
        "Review_date": 6,
        "Order Date": 7,
        "Reply Date": 8,
        "Support_reply": 9
    }
    
    stats = {
        "VERIFIED": 0,
        "RATE_LIMITED_OR_BLOCKED": 0,
        "REVIEW_NOT_FOUND": 0,
        "IDENTITY_UNVERIFIED": 0,
        "TOTAL_DISCREPANCIES": 0,
        "TOTAL_RECOVERED": 0,
        "ROWS_MODIFIED": 0
    }
    
    # Also save structured JSON report
    report_data = {}
    
    for r_idx, row_data in tasks:
        res = processed_results.get(r_idx, {"status_code": "UNKNOWN", "results": {}})
        status = res["status_code"]
        stats[status] = stats.get(status, 0) + 1
        
        ws.cell(row=r_idx, column=start_col, value=status)
        
        if status == "VERIFIED":
            field_results = res.get("results", {})
            verdicts = []
            discrepancies = []
            recovered = []
            norm_applied = []
            
            for field, f_res in field_results.items():
                v = f_res["verdict"]
                verdicts.append(f"{field}: {v}")
                if v == "DISCREPANCY":
                    discrepancies.append(field)
                elif v == "RECOVERED":
                    recovered.append(field)
                if f_res.get("norm"):
                    norm_applied.append(field)
                    
                # Update Excel cell only if discrepancy or recovered
                if v in ["DISCREPANCY", "RECOVERED"] and field in col_map:
                    target_col = col_map[field]
                    if target_col not in [3]: # Never overwrite URL
                        ws.cell(row=r_idx, column=target_col, value=f_res["final_val"])
                        
            ws.cell(row=r_idx, column=start_col+1, value=" | ".join(verdicts))
            ws.cell(row=r_idx, column=start_col+2, value=" | ".join(discrepancies))
            ws.cell(row=r_idx, column=start_col+3, value=" | ".join(recovered))
            
            made_corr = bool(discrepancies or recovered)
            if made_corr:
                stats["ROWS_MODIFIED"] += 1
                stats["TOTAL_DISCREPANCIES"] += len(discrepancies)
                stats["TOTAL_RECOVERED"] += len(recovered)
                
            ws.cell(row=r_idx, column=start_col+4, value="Yes" if made_corr else "No")
            ws.cell(row=r_idx, column=start_col+5, value="Yes" if norm_applied else "No")
            ws.cell(row=r_idx, column=start_col+6, value=f"Norms: {', '.join(norm_applied)}" if norm_applied else "Exact Match")
            
            report_data[r_idx] = {
                "status": status,
                "discrepancies": discrepancies,
                "recovered": recovered,
                "norm_applied": norm_applied
            }
        else:
            ws.cell(row=r_idx, column=start_col+6, value=res.get("reason", ""))
            
    print(f"\nSaving {TARGET_FILE}...")
    wb.save(TARGET_FILE)
    
    with open("audit_report_v2.json", "w") as f:
        json.dump(report_data, f, indent=2)
        
    print("\n" + "=" * 70)
    print("V2 RUN SUMMARY")
    print("=" * 70)
    for k, v in stats.items():
        print(f"  {k}: {v}")

if __name__ == "__main__":
    main()
