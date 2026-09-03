import requests
import json
import re
import openpyxl
from datetime import datetime
import csv
import time

INPUT_XLSM = "LG_corrected.xlsm"
OUTPUT_XLSM = "LG_audited_jina.xlsm"

def get_50_rows():
    wb = openpyxl.load_workbook(INPUT_XLSM, data_only=True)
    ws = wb["Data"]
    headers = [cell.value for cell in ws[2]]
    rows = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        row_dict = dict(zip(headers, row))
        if row_dict.get("Platform") == "Trustpilot" and row_dict.get("Url"):
            row_dict["_row_idx"] = idx
            rows.append(row_dict)
            if len(rows) == 50:
                break
    return rows

def extract_id_from_url(url):
    match = re.search(r'/reviews/([a-f0-9]+)', url)
    return match.group(1) if match else None

def test_jina_direct(url):
    jina_url = f"https://r.jina.ai/{url}"
    headers = {
        "Accept": "application/json"
    }
    try:
        resp = requests.get(jina_url, headers=headers, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            title = data.get("data", {}).get("title", "")
            content = data.get("data", {}).get("content", "")
            
            if "Verifying your connection" in title or "Verifying your connection" in content:
                return False, "WAF_BLOCKED", None
            
            if "404" in title or "Page not found" in title:
                return False, "NOT_FOUND", None
                
            return True, "SUCCESS", content
        else:
            return False, f"HTTP_{resp.status_code}", None
    except Exception as e:
        return False, str(e), None

def main():
    rows = get_50_rows()
    print(f"Selected {len(rows)} rows for verification via Jina Reader.")
    
    results = []
    
    for i, row in enumerate(rows, 1):
        url = row["Url"]
        expected_id = extract_id_from_url(url)
        print(f"[{i}/50] Row {row['_row_idx']} | Target ID: {expected_id}")
        
        success, status, content = test_jina_direct(url)
        
        match_status = "NOT_FOUND"
        if success:
            expected_text = str(row.get("Raw_text", "")).strip().lower()
            content_lower = content.lower() if content else ""
            
            if len(expected_text) > 5 and expected_text[:20] in content_lower:
                match_status = "VERIFIED_MATCH"
            elif expected_id and expected_id in content_lower:
                 match_status = "VERIFIED_MATCH"
            else:
                 match_status = "AMBIGUOUS"
        elif status == "WAF_BLOCKED":
            match_status = "ACCESS_LIMITED"
            
        print(f"    => Result: {match_status} (Jina Status: {status})")
        
        results.append({
            "row_idx": row["_row_idx"],
            "url": url,
            "status": match_status,
            "jina_status": status
        })
        
        # Courtesy delay for Jina API
        time.sleep(1)
        
    print("\n--- SUMMARY ---")
    verified = sum(1 for r in results if r["status"] == "VERIFIED_MATCH")
    limited = sum(1 for r in results if r["status"] == "ACCESS_LIMITED")
    not_found = sum(1 for r in results if r["status"] == "NOT_FOUND")
    ambiguous = sum(1 for r in results if r["status"] == "AMBIGUOUS")
    
    print(f"Total tested: {len(results)}")
    print(f"Verified: {verified}")
    print(f"Access Limited: {limited}")
    print(f"Not Found: {not_found}")
    print(f"Ambiguous: {ambiguous}")
    
    print("\nWriting to LG_audited_jina.xlsm...")
    wb = openpyxl.load_workbook(INPUT_XLSM, keep_vba=True)
    ws = wb["Data"]
    
    audit_cols = ["Jina_Pilot_Status", "Jina_Status"]
    max_col = ws.max_column
    for i, col_name in enumerate(audit_cols, start=1):
        ws.cell(row=2, column=max_col + i, value=col_name)
        
    for r in results:
        row_idx = r["row_idx"]
        ws.cell(row=row_idx, column=max_col + 1, value=r["status"])
        ws.cell(row=row_idx, column=max_col + 2, value=r["jina_status"])
        
    wb.save(OUTPUT_XLSM)
    print("Done! XLSM saved successfully.")

if __name__ == "__main__":
    main()
