import json
import time
import os
import random
import re
import csv
from datetime import datetime
from urllib.parse import urlparse
import openpyxl
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright

INPUT_XLSM = "LG_corrected.xlsm"
OUTPUT_XLSM = "LG_audited_pilot.xlsm"
CSV_LOG = "pilot_results.csv"
SUMMARY_TXT = "pilot_summary.txt"

BUSINESS_DOMAIN = "luluandgeorgia.com"

STATUSES = [
    "VERIFIED_MATCH",
    "VERIFIED_AND_CORRECTED",
    "REVIEW_NOT_FOUND",
    "MULTIPLE_POSSIBLE_MATCHES",
    "ACCESS_LIMITED",
    "SOURCE_DATA_INSUFFICIENT"
]

def load_pilot_rows(filepath, sample_size=50):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Data"]
    
    headers = [cell.value for cell in ws[2]]
    rows = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        row_dict = dict(zip(headers, row))
        if row_dict.get("Platform") == "Trustpilot":
            row_dict["_row_idx"] = idx
            rows.append(row_dict)
            
    # Stratified sampling to get a mix
    # We want a mix of ratings, dates, text lengths, and URL presence
    random.seed(42)
    random.shuffle(rows)
    
    selected = []
    
    # Try to grab diverse rows
    for r in rows:
        if len(selected) >= sample_size:
            break
        # Just grab them for now, random shuffle ensures a decent mix of the 2200 rows
        selected.append(r)
        
    return selected, headers

def extract_review_id_from_url(url):
    if not url or not isinstance(url, str):
        return None
    match = re.search(r'/reviews/([a-f0-9]+)', url)
    if match:
        return match.group(1)
    return None

def extract_reviews_from_html(html):
    soup = BeautifulSoup(html, "html.parser")
    reviews = []
    for article in soup.find_all("article"):
        r_id = article.get("id") or article.get("data-review-id")
        if not r_id:
            child = article.find(id=True)
            if child:
                r_id = child.get("id")
                
        text_el = article.find(attrs={"data-service-review-text-typography": "true"})
        text = text_el.get_text(separator=" ", strip=True) if text_el else ""
        
        rating_el = article.find(attrs={"data-service-review-rating": True})
        rating = None
        if rating_el:
            try:
                rating = int(rating_el.get("data-service-review-rating"))
            except:
                pass
                
        date_el = article.find("time")
        date_str = date_el.get("datetime") if date_el else ""
        
        reviews.append({
            "id": r_id,
            "text": text,
            "rating": rating,
            "date": date_str
        })
    return reviews

def check_waf_blocked(html):
    if "Verifying your connection" in html or "challenge.js" in html or "Verifying Connection" in page.title():
        return True
    return False

def test_retrieval_strategies(page, row):
    results = []
    
    target_id = extract_review_id_from_url(row.get("Url"))
    target_rating = row.get("Rating")
    try:
        target_rating = int(float(target_rating)) if target_rating else None
    except:
        target_rating = None
        
    strategies = []
    
    if row.get("Url") and isinstance(row.get("Url"), str) and "trustpilot.com" in row.get("Url"):
        strategies.append(("DIRECT_URL", row.get("Url")))
        
    strategies.append(("PAGINATION", f"https://www.trustpilot.com/review/{BUSINESS_DOMAIN}?page=1"))
    
    if target_rating:
        strategies.append(("STAR_FILTER", f"https://www.trustpilot.com/review/{BUSINESS_DOMAIN}?stars={target_rating}"))
        
    strategies.append(("SORTING", f"https://www.trustpilot.com/review/{BUSINESS_DOMAIN}?sort=recency"))

    match_status = "REVIEW_NOT_FOUND"
    matched_review = None
    successful_strategy = None
    access_blocked = False
    
    for strategy_name, url in strategies:
        print(f"    -> Attempting {strategy_name}: {url}")
        
        try:
            resp = page.goto(url, wait_until="domcontentloaded", timeout=15000)
            page.wait_for_timeout(2000) # Give it a moment to render or trigger WAF
            
            html = page.content()
            
            # Check WAF
            if check_waf_blocked(html):
                results.append({
                    "strategy": strategy_name,
                    "url": url,
                    "status": resp.status if resp else None,
                    "waf_blocked": True,
                    "reviews_found": 0,
                    "matched": False
                })
                access_blocked = True
                continue
            
            # Extracted reviews
            page_reviews = extract_reviews_from_html(html)
            
            results.append({
                "strategy": strategy_name,
                "url": url,
                "status": resp.status if resp else None,
                "waf_blocked": False,
                "reviews_found": len(page_reviews),
                "matched": False
            })
            
            # Check match
            match_found = False
            candidates = []
            
            for rev in page_reviews:
                # 1. Exact ID match
                if target_id and rev["id"] == target_id:
                    match_found = True
                    matched_review = rev
                    break
                    
                # 2. Composite match (Text + Rating)
                row_text = str(row.get("Raw_text", "")).strip().lower()
                rev_text = str(rev.get("text", "")).strip().lower()
                
                # Check if text is remarkably similar and rating matches
                if len(row_text) > 10 and row_text in rev_text and target_rating == rev.get("rating"):
                    candidates.append(rev)
            
            if not match_found and len(candidates) == 1:
                match_found = True
                matched_review = candidates[0]
            elif not match_found and len(candidates) > 1:
                match_status = "MULTIPLE_POSSIBLE_MATCHES"
                break
                
            if match_found:
                match_status = "VERIFIED_MATCH"
                successful_strategy = strategy_name
                results[-1]["matched"] = True
                break
                
        except Exception as e:
            print(f"       [Error] {str(e)}")
            results.append({
                "strategy": strategy_name,
                "url": url,
                "status": None,
                "waf_blocked": False,
                "reviews_found": 0,
                "matched": False,
                "error": str(e)
            })

    # If we never matched but we hit a WAF on any strategy, default to access limited unless multiple found
    if match_status == "REVIEW_NOT_FOUND" and access_blocked:
        match_status = "ACCESS_LIMITED"
        
    if not row.get("Url") and not row.get("Raw_text"):
        match_status = "SOURCE_DATA_INSUFFICIENT"
        
    return match_status, matched_review, successful_strategy, results

def main():
    print("Loading pilot rows...")
    pilot_rows, headers = load_pilot_rows(INPUT_XLSM, 50)
    print(f"Selected {len(pilot_rows)} rows for pilot.")
    
    results_log = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
        )
        
        for idx, row in enumerate(pilot_rows, 1):
            print(f"\n[{idx}/50] Row {row['_row_idx']} | Rating: {row.get('Rating')} | URL: {row.get('Url')}")
            page = context.new_page()
            
            match_status, matched_review, successful_strategy, attempts = test_retrieval_strategies(page, row)
            
            print(f"    => Status: {match_status} | Strategy: {successful_strategy}")
            
            # Save for aggregate
            results_log.append({
                "row_id": row["_row_idx"],
                "spreadsheet_review_date": row.get("Review_date"),
                "spreadsheet_rating": row.get("Rating"),
                "spreadsheet_text": row.get("Raw_text"),
                "spreadsheet_url": row.get("Url"),
                "verification_status": match_status,
                "successful_strategy": successful_strategy,
                "matched_review_id": matched_review["id"] if matched_review else None,
                "attempts": json.dumps(attempts)
            })
            
            page.close()
            time.sleep(1) # Courtesy delay
            
        browser.close()

    # Write CSV Log
    with open(CSV_LOG, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=results_log[0].keys())
        writer.writeheader()
        writer.writerows(results_log)
        
    # Generate Summary
    stats = {
        "Total pilot rows": 50,
        "Verified": 0,
        "Verified + corrected": 0,
        "Not found": 0,
        "Ambiguous": 0,
        "Access limited": 0,
        "Source data insufficient": 0
    }
    
    strategies_stats = {
        "DIRECT_URL": 0,
        "PAGINATION": 0,
        "STAR_FILTER": 0,
        "SORTING": 0
    }
    
    for r in results_log:
        s = r["verification_status"]
        if s == "VERIFIED_MATCH":
            stats["Verified"] += 1
            if r["successful_strategy"]:
                strategies_stats[r["successful_strategy"]] += 1
        elif s == "VERIFIED_AND_CORRECTED":
            stats["Verified + corrected"] += 1
        elif s == "REVIEW_NOT_FOUND":
            stats["Not found"] += 1
        elif s == "MULTIPLE_POSSIBLE_MATCHES":
            stats["Ambiguous"] += 1
        elif s == "ACCESS_LIMITED":
            stats["Access limited"] += 1
        elif s == "SOURCE_DATA_INSUFFICIENT":
            stats["Source data insufficient"] += 1
            
    with open(SUMMARY_TXT, "w") as f:
        f.write("ROW-LEVEL SUMMARY\n")
        f.write("=================\n")
        for r in results_log:
            f.write(f"Row: {r['row_id']} | Status: {r['verification_status']} | Strategy: {r['successful_strategy']} | ID: {r['matched_review_id']}\n")
            
        f.write("\nAGGREGATE SUMMARY\n")
        f.write("=================\n")
        for k, v in stats.items():
            f.write(f"{k}: {v}\n")
            
        f.write("\nRetrieval strategy success:\n")
        for k, v in strategies_stats.items():
            f.write(f"{k}: {v}\n")
            
        verified_count = stats['Verified'] + stats['Verified + corrected']
        f.write(f"\nOverall verification rate: {verified_count / 50.0:.1%}\n")
        
        non_access_limited = 50 - stats["Access limited"]
        if non_access_limited > 0:
            f.write(f"Verification rate excluding access-limited rows: {verified_count / non_access_limited:.1%}\n")
        f.write(f"Access-limited rate: {stats['Access limited'] / 50.0:.1%}\n")
        
    # Write XLSM Pilot Output
    print("\nWriting output XLSM...")
    wb = openpyxl.load_workbook(INPUT_XLSM, keep_vba=True)
    ws = wb["Data"]
    
    # Append Audit columns to header
    audit_cols = ["Pilot_Status", "Pilot_Strategy", "Pilot_Matched_ID"]
    max_col = ws.max_column
    for i, col_name in enumerate(audit_cols, start=1):
        ws.cell(row=2, column=max_col + i, value=col_name)
        
    for r in results_log:
        row_idx = r["row_id"]
        ws.cell(row=row_idx, column=max_col + 1, value=r["verification_status"])
        ws.cell(row=row_idx, column=max_col + 2, value=r["successful_strategy"])
        ws.cell(row=row_idx, column=max_col + 3, value=r["matched_review_id"])
        
    wb.save(OUTPUT_XLSM)
    print(f"Pilot execution finished. Created {CSV_LOG}, {SUMMARY_TXT}, and {OUTPUT_XLSM}.")

if __name__ == "__main__":
    main()
