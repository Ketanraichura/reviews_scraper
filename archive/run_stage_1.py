import requests
import json
import re
import openpyxl
from bs4 import BeautifulSoup
import time

INPUT_XLSM = "LG_corrected.xlsm"
AMBIGUOUS_ROWS = [12, 13, 19, 40, 49] # Indices of ambiguous rows from the 50-row test list

def get_regression_rows():
    wb = openpyxl.load_workbook(INPUT_XLSM, data_only=True)
    ws = wb["Data"]
    headers = [cell.value for cell in ws[2]]
    rows = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        row_dict = dict(zip(headers, row))
        if row_dict.get("Platform") == "Trustpilot" and row_dict.get("Url"):
            row_dict["_row_idx"] = idx
            rows.append(row_dict)
            if len(rows) == 55:
                break
                
    targets = []
    for i in AMBIGUOUS_ROWS:
        targets.append(rows[i - 1])
    return targets

def extract_id_from_url(url):
    match = re.search(r'/reviews/([a-f0-9]+)', url)
    return match.group(1) if match else None

def extract_review_from_next_data(html_content, expected_id):
    """Parses __NEXT_DATA__ JSON from HTML and extracts the review."""
    soup = BeautifulSoup(html_content, "html.parser")
    script = soup.find("script", id="__NEXT_DATA__")
    
    if not script:
        return False, "MISSING_NEXT_DATA", None
        
    try:
        data = json.loads(script.string)
    except json.JSONDecodeError:
        return False, "MALFORMED_JSON", None
        
    # Trustpilot single review pages store data in props.pageProps.review
    try:
        page_props = data.get("props", {}).get("pageProps", {})
        review_obj = page_props.get("review")
        
        if review_obj and review_obj.get("id") == expected_id:
            extracted_data = {
                "id": review_obj.get("id"),
                "text": review_obj.get("text"),
                "rating": review_obj.get("rating"),
                "date": review_obj.get("dates", {}).get("experiencedDate") or review_obj.get("dates", {}).get("publishedDate")
            }
            return True, "SUCCESS", extracted_data
            
        # Fallback to ApolloState if it's there
        apollo_state = page_props.get("apolloState", {})
        review_key = f"Review:{expected_id}"
        if review_key in apollo_state:
            val = apollo_state[review_key]
            extracted_data = {
                "id": val.get("id"),
                "text": val.get("text"),
                "rating": val.get("rating"),
                "date": val.get("dates", {}).get("experiencedDate") or val.get("dates", {}).get("publishedDate")
            }
            return True, "SUCCESS", extracted_data
                
        return False, "REVIEW_ID_NOT_IN_PAYLOAD", None
    except Exception as e:
        return False, f"PARSING_ERROR: {str(e)}", None

def test_jina_html(url, expected_id):
    jina_url = f"https://r.jina.ai/{url}"
    headers = {
        "Accept": "application/json",
        "X-Return-Format": "html"
    }
    try:
        resp = requests.get(jina_url, headers=headers, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            title = data.get("data", {}).get("title", "")
            html_content = data.get("data", {}).get("html", "")
            
            if "Verifying" in title or "Verifying Connection" in title or "challenge.js" in html_content:
                return False, "ACCESS_LIMITED", None
            
            if "404" in title or "Page not found" in title:
                return False, "REVIEW_NOT_FOUND", None
                
            return extract_review_from_next_data(html_content, expected_id)
        else:
            return False, f"HTTP_{resp.status_code}", None
    except Exception as e:
        return False, str(e), None

def run_stage_1():
    rows = get_regression_rows()
    print(f"Stage 1: Testing {len(rows)} ambiguous rows with __NEXT_DATA__ extractor.")
    
    results = []
    
    for i, row in enumerate(rows, 1):
        url = row["Url"]
        expected_id = extract_id_from_url(url)
        print(f"[{i}/5] Row {row['_row_idx']} | Target ID: {expected_id}")
        
        success, status, review_data = test_jina_html(url, expected_id)
        
        match_status = "SOURCE_DATA_INSUFFICIENT"
        
        if success and review_data:
            # Identity is unequivocally proven if the structured payload yields the exact Review ID
            if review_data.get("id") == expected_id:
                match_status = "VERIFIED_MATCH"
            else:
                match_status = "MULTIPLE_POSSIBLE_MATCHES" # the new AMBIGUOUS
        elif status == "ACCESS_LIMITED":
            match_status = "ACCESS_LIMITED"
        elif status == "REVIEW_NOT_FOUND":
            match_status = "REVIEW_NOT_FOUND"
        else:
            # Fallback if next_data fails
            print(f"    Extractor failure: {status}")
            
        print(f"    => Result: {match_status} (Extraction Status: {status})")
        
        results.append({
            "row_idx": row["_row_idx"],
            "url": url,
            "status": match_status,
        })
        time.sleep(2)
        
    print("\n--- STAGE 1 SUMMARY ---")
    verified = sum(1 for r in results if r["status"] == "VERIFIED_MATCH")
    
    if verified == len(rows):
        print(f"SUCCESS! 5/5 regression rows flipped to VERIFIED_MATCH.")
    else:
        print(f"FAILURE. Only {verified}/5 became VERIFIED_MATCH.")

if __name__ == "__main__":
    run_stage_1()
