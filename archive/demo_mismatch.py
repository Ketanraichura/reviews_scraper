import openpyxl
import re
import string
import requests
import json

INPUT_XLSM = "LG_corrected.xlsm"

def extract_id_from_url(url):
    match = re.search(r'/reviews/([a-f0-9]+)', str(url))
    return match.group(1) if match else None

def normalize_text(text):
    if not text: return ""
    text = str(text).lower()
    text = text.translate(str.maketrans('', '', string.punctuation))
    return re.sub(r'\s+', ' ', text).strip()

def check_identity(expected_id, expected_text, content):
    content_lower = content.lower() if content else ""
    if expected_id and expected_id in content_lower:
        return True, "ID_MATCH"
    norm_expected = normalize_text(expected_text)
    norm_content = normalize_text(content)
    if not norm_expected: return False, "NO_EXPECTED_TEXT"
    if norm_expected in norm_content: return True, "EXACT_TEXT_MATCH"
    sentences = re.split(r'[.!?]+', str(expected_text))
    if sentences:
        longest_sentence = max(sentences, key=len).strip()
        norm_longest = normalize_text(longest_sentence)
        if len(norm_longest) > 20 and norm_longest in norm_content:
            return True, "STRONG_PARTIAL_TEXT_MATCH"
    return False, "INSUFFICIENT_EVIDENCE"

def extract_fields(content):
    lines = content.split('\n')
    text_lines = []
    review_date = None
    reply_date = None
    reply_lines = []
    
    in_reply = False
    date_pattern = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},\s+\d{4}$", re.IGNORECASE)
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith("## [") or line.startswith("![User"):
            continue
            
        if "**Reply from" in line or "Reply from" in line:
            in_reply = True
            continue
            
        if date_pattern.match(line):
            if in_reply and not reply_date:
                reply_date = line
            elif not in_reply and not review_date:
                review_date = line
            continue
            
        if in_reply:
            reply_lines.append(line)
        else:
            text_lines.append(line)
            
    return {
        "Raw_text": "\n".join(text_lines).strip() if text_lines else None,
        "Review_date": review_date,
        "Support_reply": "\n".join(reply_lines).strip() if reply_lines else None,
        "Reply_date": reply_date,
        "Rating": None
    }

def run_demo():
    print("--- DEMONSTRATION OF FIELD CORRECTION ---")
    
    # 1. Load Excel row (Row 18)
    wb = openpyxl.load_workbook(INPUT_XLSM, data_only=True)
    ws = wb["Data"]
    headers = [cell.value for cell in ws[2]]
    
    row_idx = 18
    row_values = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    row = dict(zip(headers, row_values))
    
    url = row["Url"]
    expected_id = extract_id_from_url(url)
    
    # 2. Deliberately Mismatch the Excel Fields
    excel_before = {
        "Raw_text": "THIS IS DELIBERATELY FAKE TEXT TO PROVE MISMATCH HANDLING.",
        "Review_date": "1999-01-01",
        "Rating": 1
    }
    
    print(f"\nTARGET URL: {url}")
    print("\n[EXCEL BEFORE]")
    for k, v in excel_before.items():
        print(f"{k}: {v}")
        
    # 3. Fetch from Trustpilot
    print("\nFetching from Trustpilot (via Jina Markdown)...")
    jina_url = f"https://r.jina.ai/{url}"
    resp = requests.get(jina_url, headers={"Accept": "application/json"}, timeout=30)
    data = resp.json()
    content = data.get("data", {}).get("content", "")
    
    # 4. Verify Identity
    is_match, reason = check_identity(expected_id, row["Raw_text"], content)
    print(f"Identity Verification: {'SUCCESS' if is_match else 'FAILED'} ({reason})")
    
    if not is_match:
        print("Cannot proceed with correction if identity is not verified.")
        return
        
    # 5. Extract Trustpilot Data
    tp_fields = extract_fields(content)
    print("\n[TRUSTPILOT EXTRACTED DATA]")
    for k, v in tp_fields.items():
        # Truncate text for display if long
        display_v = (str(v)[:100] + "...") if v and len(str(v)) > 100 else v
        print(f"{k}: {display_v}")
        
    # 6. Compare and Correct
    print("\n[CORRECTION PROCESS]")
    audit_records = []
    excel_after = excel_before.copy()
    
    for field in ["Raw_text", "Review_date", "Rating", "Reply_date", "Support_reply"]:
        orig_val = excel_before.get(field)
        tp_val = tp_fields.get(field)
        
        if tp_val is not None:
            # Normalize strings for comparison
            orig_str = str(orig_val).strip().replace('\r\n', '\n') if orig_val else ""
            tp_str = str(tp_val).strip()
            
            if orig_str != tp_str:
                print(f"Mismatch in {field}! Correcting...")
                excel_after[field] = tp_val
                audit_records.append({
                    "Verification_Status": "VERIFIED_AND_CORRECTED",
                    "Field_Affected": field,
                    "Original_Value": str(orig_val)[:100],
                    "Trustpilot_Value": str(tp_val)[:100],
                    "Correction_Made": "Yes",
                    "Evidence_Source": "Jina Markdown Extract",
                    "Confidence": "High",
                    "Reason_Action_Taken": f"Reliable Trustpilot value differed from Excel."
                })
            else:
                print(f"{field} matches perfectly. No correction needed.")
        else:
            print(f"{field} not reliably extracted from Trustpilot. Leaving Excel unchanged.")
            
    print("\n[EXCEL AFTER]")
    for k, v in excel_after.items():
        # Truncate text for display if long
        display_v = (str(v)[:100] + "...") if v and len(str(v)) > 100 else v
        print(f"{k}: {display_v}")
        
    print("\n[AUDIT RECORDS WRITTEN]")
    for r in audit_records:
        print(json.dumps(r, indent=2))

if __name__ == "__main__":
    run_demo()
