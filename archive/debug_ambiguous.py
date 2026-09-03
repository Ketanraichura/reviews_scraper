import requests
import json
import openpyxl

INPUT_XLSM = "LG_corrected.xlsm"

AMBIGUOUS_ROWS = [15, 16, 22, 43, 52] # Index of the row inside our get_50_rows logic

def get_ambiguous_urls():
    wb = openpyxl.load_workbook(INPUT_XLSM, data_only=True)
    ws = wb["Data"]
    headers = [cell.value for cell in ws[2]]
    rows = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        row_dict = dict(zip(headers, row))
        if row_dict.get("Platform") == "Trustpilot" and row_dict.get("Url"):
            row_dict["_row_idx"] = idx
            rows.append(row_dict)
            if len(rows) == 55:
                break
                
    # Get the 0-indexed items matching the 1-indexed log output
    targets = []
    for i in AMBIGUOUS_ROWS:
        targets.append(rows[i - 1]) # Convert 1-indexed to 0-indexed
        
    return targets

def fetch_and_dump(row):
    url = row["Url"]
    print(f"Fetching: {url}")
    jina_url = f"https://r.jina.ai/{url}"
    headers = {"Accept": "application/json"}
    
    resp = requests.get(jina_url, headers=headers)
    if resp.status_code == 200:
        data = resp.json()
        content = data.get("data", {}).get("content", "")
        
        with open(f"debug_row_{row['_row_idx']}.md", "w") as f:
            f.write(content)
            
        print(f"-> Saved to debug_row_{row['_row_idx']}.md")
    else:
        print(f"-> Failed: {resp.status_code}")

def main():
    targets = get_ambiguous_urls()
    for row in targets:
        fetch_and_dump(row)

if __name__ == "__main__":
    main()
