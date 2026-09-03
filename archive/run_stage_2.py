import requests
import json
import re
import openpyxl
import time
import string

INPUT_XLSM = "LG_corrected.xlsm"

def get_50_rows():
    wb = openpyxl.load_workbook(INPUT_XLSM, data_only=True)
    ws = wb["Data"]
    headers = [cell.value for cell in ws[2]]
    rows = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        row_dict = dict(zip(headers, row))
        if row_dict.get("Platform") == "Trustpilot" and row_dict.get("Url"):
            row_dict["_row_idx"] = idx
            rows.append(row_dict)
            if len(rows) == 50:
                break
    return rows

def extract_id_from_url(url):
    match = re.search(r'/reviews/([a-f0-9]+)', url)
    return match.group(1) if match else None

def normalize_text(text):
    if not text:
        return ""
    text = str(text).lower()
    text = text.translate(str.maketrans('', '', string.punctuation))
    return re.sub(r'\s+', ' ', text).strip()

def check_identity(expected_id, expected_text, content):
    content_lower = content.lower() if content else ""
    
    if expected_id and expected_id in content_lower:
        return True, "ID_MATCH"
        
    norm_expected = normalize_text(expected_text)
    norm_content = normalize_text(content)
    
    if not norm_expected:
        return False, "NO_EXPECTED_TEXT"
        
    if norm_expected in norm_content:
        return True, "EXACT_TEXT_MATCH"
        
    sentences = re.split(r'[.!?]+', str(expected_text))
    longest_sentence = max(sentences, key=len).strip()
    norm_longest = normalize_text(longest_sentence)
    
    if len(norm_longest) > 20 and norm_longest in norm_content:
        return True, "STRONG_PARTIAL_TEXT_MATCH"
        
    return False, "INSUFFICIENT_EVIDENCE"

def test_jina_markdown(url, expected_id, expected_text):
    jina_url = f"https://r.jina.ai/{url}"
    headers = {
        "Accept": "application/json"
    }
    try:
        resp = requests.get(jina_url, headers=headers, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            title = data.get("data", {}).get("title", "")
            content = data.get("data", {}).get("content", "")
            
            if "Verifying your connection" in title or "Verifying" in title or "Verifying Connection" in title:
                return False, "ACCESS_LIMITED", None
            
            if "404" in title or "Page not found" in title:
                return False, "REVIEW_NOT_FOUND", None
                
            is_match, reason = check_identity(expected_id, expected_text, content)
            
            if is_match:
                return True, "VERIFIED_MATCH", reason
            else:
                return False, "SOURCE_DATA_INSUFFICIENT", reason
                
        else:
            return False, f"HTTP_{resp.status_code}", None
    except Exception as e:
        return False, str(e), None

def run_stage_2():
    rows = get_50_rows()
    print(f"Stage 2: Testing {len(rows)} pilot rows with Improved Markdown fallback.")
    
    results = []
    
    for i, row in enumerate(rows, 1):
        url = row["Url"]
        expected_id = extract_id_from_url(url)
        expected_text = row.get("Raw_text", "")
        
        success, status, reason = test_jina_markdown(url, expected_id, expected_text)
        
        if success:
            match_status = "VERIFIED_MATCH"
        else:
            match_status = status
            
        print(f"[{i}/50] Row {row['_row_idx']} | => Result: {match_status} (Reason: {reason})")
        
        results.append({
            "row_idx": row["_row_idx"],
            "url": url,
            "status": match_status,
        })
        time.sleep(2)
        
    print("\n--- STAGE 2 SUMMARY ---")
    verified = sum(1 for r in results if r["status"] == "VERIFIED_MATCH")
    limited = sum(1 for r in results if r["status"] == "ACCESS_LIMITED")
    not_found = sum(1 for r in results if r["status"] == "REVIEW_NOT_FOUND")
    ambiguous = sum(1 for r in results if r["status"] == "SOURCE_DATA_INSUFFICIENT")
    
    print(f"Total tested: {len(results)}")
    print(f"VERIFIED_MATCH: {verified}")
    print(f"SOURCE_DATA_INSUFFICIENT: {ambiguous}")
    print(f"REVIEW_NOT_FOUND: {not_found}")
    print(f"ACCESS_LIMITED: {limited}")
    
if __name__ == "__main__":
    run_stage_2()
