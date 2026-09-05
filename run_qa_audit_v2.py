import os
import json
import glob
import openpyxl

from src_v2.extractor import extract_all_fields
from src_v2.comparator import compare_all

SOURCE_FILE = "LG_corrected.xlsm"
CACHE_DIR = "cache/raw_reviews"

def main():
    print("=" * 70)
    print("V2 QA AUDIT RUNNER")
    print("=" * 70)
    
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    ws = wb["Data"]
    
    headers = [ws.cell(row=2, column=c).value for c in range(1, 10)]
    rows_by_id = {}
    
    for r in range(4, ws.max_row + 1):
        url = ws.cell(row=r, column=3).value
        if url:
            import re
            m = re.search(r"/reviews/([a-f0-9]+)", str(url))
            if m:
                rows_by_id[m.group(1)] = {headers[c-1]: ws.cell(row=r, column=c).value for c in range(1, 10)}
                
    files = glob.glob(f"{CACHE_DIR}/*.json")
    print(f"Running QA on {len(files)} cached JSON retrievals...\n")
    
    for fp in files:
        rev_id = os.path.basename(fp).replace(".json", "")
        row_data = rows_by_id.get(rev_id)
        if not row_data:
            continue
            
        with open(fp) as f:
            d = json.load(f)
            
        content = d.get("data", {}).get("content", "")
        title = d.get("data", {}).get("title", "")
        
        expected_text = row_data.get("Raw_text")
        
        res = extract_all_fields(content, expected_id=rev_id, expected_text=expected_text, title_meta=title)
        
        if not res["Identity_Confirmed"]:
            print(f"[{rev_id}] IDENTITY UNVERIFIED: {res.get('Reason')}")
            continue
            
        comp = compare_all(row_data, res)
        
        discrepancies = []
        recovered = []
        norms = []
        
        for field, f_res in comp.items():
            if f_res["verdict"] == "DISCREPANCY":
                discrepancies.append(f"{field} ('{f_res['orig_val']}' -> '{f_res['tp_val']}')")
            elif f_res["verdict"] == "RECOVERED":
                recovered.append(f"{field} (None -> '{f_res['tp_val']}')")
            elif f_res.get("norm"):
                norms.append(f"{field} ('{f_res['orig_val']}' == '{f_res['tp_val']}')")
                
        print(f"--- Review {rev_id} ---")
        if discrepancies:
            print("  DISCREPANCIES:")
            for d_str in discrepancies:
                print(f"    - {d_str}")
        if recovered:
            print("  RECOVERED:")
            for r_str in recovered:
                print(f"    + {r_str}")
        if norms:
            print("  NORMALIZED (MATCH):")
            for n_str in norms:
                print(f"    ~ {n_str}")
        if not discrepancies and not recovered and not norms:
            print("  PERFECT MATCH")
            
    print("\nQA Audit Complete.")

if __name__ == "__main__":
    main()
