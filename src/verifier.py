import requests
import re
import string
import time
import openpyxl
import json
import os
import threading
import random
import concurrent.futures
from collections import deque
from copy import deepcopy

class TrustpilotVerifier:
    def __init__(self, input_file, output_file):
        self.input_file = input_file
        self.output_file = output_file
        
    def _normalize_text(self, text):
        if not text: return ""
        text = str(text).lower()
        text = text.translate(str.maketrans('', '', string.punctuation))
        return re.sub(r'\s+', ' ', text).strip()
        
    def _extract_id_from_url(self, url):
        match = re.search(r'/reviews/([a-f0-9]+)', str(url))
        return match.group(1) if match else None

    def _fetch_jina_markdown(self, url):
        jina_url = f"https://r.jina.ai/{url}"
        headers = {"Accept": "application/json"}
        try:
            resp = requests.get(jina_url, headers=headers, timeout=30)
            if resp.status_code == 200:
                data = resp.json()
                title = data.get("data", {}).get("title", "")
                content = data.get("data", {}).get("content", "")
                
                if "Verifying your connection" in title or "Verifying" in title or "Verifying Connection" in title:
                    return False, "ACCESS_LIMITED", None
                if "404" in title or "Page not found" in title:
                    return False, "REVIEW_NOT_FOUND", None
                    
                return True, "SUCCESS", content
            return False, f"HTTP_{resp.status_code}", None
        except Exception as e:
            return False, str(e), None

    def _verify_identity(self, expected_id, expected_text, content):
        content_lower = content.lower() if content else ""
        if expected_id and expected_id in content_lower:
            return True, "ID_MATCH"
            
        norm_expected = self._normalize_text(expected_text)
        norm_content = self._normalize_text(content)
        
        if not norm_expected:
            return False, "NO_EXPECTED_TEXT"
            
        if norm_expected in norm_content:
            return True, "EXACT_TEXT_MATCH"
            
        sentences = re.split(r'[.!?]+', str(expected_text))
        if sentences:
            longest_sentence = max(sentences, key=len).strip()
            norm_longest = self._normalize_text(longest_sentence)
            if len(norm_longest) > 20 and norm_longest in norm_content:
                return True, "STRONG_PARTIAL_TEXT_MATCH"
                
        return False, "INSUFFICIENT_EVIDENCE"

    def _safely_extract_fields(self, content):
        lines = content.split('\n')
        text_lines = []
        review_date = None
        reply_date = None
        reply_lines = []
        
        in_review_body = False
        in_reply = False
        
        date_pattern = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},\s+\d{4}$", re.IGNORECASE)
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if line.startswith("## ["):
                in_review_body = True
                continue
                
            if "**Reply from" in line or "Reply from" in line:
                in_review_body = False
                in_reply = True
                continue
                
            if in_review_body:
                if date_pattern.match(line):
                    review_date = line
                    in_review_body = False
                    continue
                if line == "* * *" or line == "Useful" or line == "Share":
                    in_review_body = False
                    continue
                    
                text_lines.append(line)
                
            elif in_reply:
                if date_pattern.match(line) and not reply_date:
                    reply_date = line
                    continue
                if line == "* * *":
                    in_reply = False
                    continue
                    
                reply_lines.append(line)
                
            else:
                if date_pattern.match(line) and not review_date:
                    review_date = line

        tp_text = "\n".join(text_lines).strip() if text_lines else None
        
        # Safeguard: If the extracted text has markdown artifacts, reject it
        if tp_text and ("![Image" in tp_text or "](https" in tp_text):
            tp_text = None

        return {
            "Raw_text": tp_text,
            "Review_date": review_date,
            "Support_reply": "\n".join(reply_lines).strip() if reply_lines else None,
            "Reply_date": reply_date,
            "Rating": None
        }

    def _process_single_row(self, r_data):
        """Pure function to process a single row safely in a thread."""
        row_idx = r_data["_row_idx"]
        url = r_data["Url"]
        expected_id = self._extract_id_from_url(url)
        expected_text = r_data.get("Raw_text", "")
        
        max_retries = 4
        success = False
        status = "UNKNOWN"
        content = None
        http_code = 0
        latency = 0
        
        for attempt in range(max_retries):
            jina_url = f"https://r.jina.ai/{url}"
            try:
                start_time = time.time()
                resp = requests.get(jina_url, headers={"Accept": "application/json"}, timeout=30)
                latency = time.time() - start_time
                http_code = resp.status_code
                
                if resp.status_code == 429:
                    status = "RATE_LIMITED"
                elif resp.status_code == 200:
                    data = resp.json()
                    title = data.get("data", {}).get("title", "")
                    content = data.get("data", {}).get("content", "")
                    
                    if "Verifying your connection" in title or "Verifying" in title or "challenge.js" in content or "Verifying Connection" in title:
                        status = "ACCESS_LIMITED"
                    elif "404" in title or "Page not found" in title:
                        status = "REVIEW_NOT_FOUND"
                    else:
                        success = True
                        status = "SUCCESS"
                        break
                else:
                    status = f"HTTP_{resp.status_code}"
            except Exception as e:
                status = f"ERROR_{str(e)[:20]}"
                latency = time.time() - start_time
                
            if status in ["ACCESS_LIMITED", "RATE_LIMITED"] or "ERROR" in status:
                sleep_time = (2 ** attempt) + random.uniform(0, 1)
                time.sleep(sleep_time)
            else:
                break
                
        status_code = "SOURCE_DATA_INSUFFICIENT"
        audit_records = []
        corrections = {}
        
        if success:
            is_match, reason = self._verify_identity(expected_id, expected_text, content)
            if is_match:
                status_code = "VERIFIED_MATCH"
                tp_fields = self._safely_extract_fields(content)
                
                col_map = {
                    "Raw_text": 3,
                    "Rating": 4,
                    "Review_date": 5,
                    "Reply_date": 6,
                    "Support_reply": 7
                }
                
                row_corrected = False
                for field, col_idx in col_map.items():
                    orig_val = r_data.get(field)
                    tp_val = tp_fields.get(field)
                    
                    if tp_val is not None:
                        orig_str = str(orig_val).strip().replace('\r\n', '\n') if orig_val else ""
                        tp_str = str(tp_val).strip()
                        
                        if orig_str != tp_str:
                            # QA SAFEGUARD for Raw_text: 
                            if field == "Raw_text":
                                if tp_str.endswith("...") or tp_str.endswith("…"):
                                    continue # Truncated value
                                if len(tp_str) < (len(orig_str) * 0.5):
                                    continue # Suspiciously short, likely partial extraction
                                    
                            corrections[field] = tp_val
                            row_corrected = True
                            audit_records.append({
                                "Field_Affected": field,
                                "Original_Value": str(orig_val)[:100],
                                "Trustpilot_Value": str(tp_val)[:100],
                                "Correction_Made": "Yes",
                                "Evidence_Source": "Jina Markdown Extract",
                                "Confidence": "High",
                                "Reason_Action_Taken": "Reliable Trustpilot value differed from Excel."
                            })
                            
                if row_corrected:
                    status_code = "VERIFIED_AND_CORRECTED"
                
                if not audit_records:
                    audit_records.append({
                        "Field_Affected": "None",
                        "Original_Value": "N/A",
                        "Trustpilot_Value": "N/A",
                        "Correction_Made": "No",
                        "Evidence_Source": "Jina Markdown",
                        "Confidence": "High",
                        "Reason_Action_Taken": "Identity verified. No discrepancies found or missing fields not reliably extractable."
                    })
            else:
                status_code = "MULTIPLE_POSSIBLE_MATCHES"
                audit_records.append({
                    "Field_Affected": "None",
                    "Original_Value": "N/A",
                    "Trustpilot_Value": "N/A",
                    "Correction_Made": "No",
                    "Evidence_Source": "Jina Markdown",
                    "Confidence": "Low",
                    "Reason_Action_Taken": f"Content retrieved but identity unconfirmed. Reason: {reason}"
                })
        elif status in ["ACCESS_LIMITED", "REVIEW_NOT_FOUND", "RATE_LIMITED"]:
            status_code = status
            audit_records.append({
                "Field_Affected": "None",
                "Original_Value": "N/A",
                "Trustpilot_Value": "N/A",
                "Correction_Made": "No",
                "Evidence_Source": "HTTP Response",
                "Confidence": "High",
                "Reason_Action_Taken": f"Failed to retrieve content. Status: {status}"
            })
            
        return {
            "row_idx": row_idx,
            "status_code": status_code,
            "audit_records": audit_records,
            "corrections": corrections,
            "latency": latency,
            "http_status": status
        }

    def concurrent_verify(self, limit=None, checkpoint_file="checkpoint.json"):
        print(f"Loading {self.input_file} for data reading (data_only=True)...")
        wb_data = openpyxl.load_workbook(self.input_file, data_only=True)
        ws_data = wb_data["Data"]
        headers = [cell.value for cell in ws_data[2]]
        
        target_rows = []
        for idx, row in enumerate(ws_data.iter_rows(min_row=3, values_only=True), start=3):
            row_dict = dict(zip(headers, row))
            if row_dict.get("Platform") == "Trustpilot" and row_dict.get("Url"):
                row_dict["_row_idx"] = idx
                target_rows.append(row_dict)
                
        if limit:
            target_rows = target_rows[:limit]
            
        completed = {}
        if os.path.exists(checkpoint_file):
            with open(checkpoint_file, 'r') as f:
                completed = json.load(f)
                
        pending_rows = [r for r in target_rows if str(r["_row_idx"]) not in completed]
        
        print(f"Total target rows: {len(target_rows)}")
        print(f"Already completed: {len(completed)}")
        print(f"Pending processing: {len(pending_rows)}")
        
        current_concurrency = 2
        max_concurrency = 5
        min_concurrency = 1
        
        success_streak = 0
        error_streak = 0
        
        results_queue = []
        latencies = deque(maxlen=100)
        
        start_time = time.time()
        total_processed_session = 0
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_concurrency) as executor:
            futures = {}
            pending_queue = deque(pending_rows)
            
            while pending_queue or futures:
                while len(futures) < current_concurrency and pending_queue:
                    row = pending_queue.popleft()
                    future = executor.submit(self._process_single_row, row)
                    futures[future] = row["_row_idx"]
                    
                done, not_done = concurrent.futures.wait(
                    futures.keys(), return_when=concurrent.futures.FIRST_COMPLETED
                )
                
                for f in done:
                    r_idx = futures.pop(f)
                    try:
                        res = f.result()
                        results_queue.append(res)
                        latencies.append(res["latency"])
                        total_processed_session += 1
                        
                        if res["http_status"] in ["SUCCESS", "REVIEW_NOT_FOUND"]:
                            success_streak += 1
                            error_streak = 0
                            if success_streak >= 20 and current_concurrency < max_concurrency:
                                current_concurrency += 1
                                success_streak = 0
                        elif res["http_status"] in ["ACCESS_LIMITED", "RATE_LIMITED"] or "ERROR" in res["http_status"]:
                            error_streak += 1
                            success_streak = 0
                            if error_streak >= 2:
                                current_concurrency = max(min_concurrency, current_concurrency - 2)
                                error_streak = 0
                                time.sleep(5)
                                
                    except Exception as exc:
                        print(f"Row {r_idx} generated an exception: {exc}")
                        error_streak += 1
                        current_concurrency = max(min_concurrency, current_concurrency - 1)
                        
                if total_processed_session % 10 == 0 and total_processed_session > 0:
                    elapsed = time.time() - start_time
                    rate = total_processed_session / elapsed if elapsed > 0 else 0
                    rem = len(pending_queue) + len(futures)
                    eta = rem / rate if rate > 0 else 0
                    avg_lat = sum(latencies)/len(latencies) if latencies else 0
                    if latencies:
                        sorted_lats = sorted(list(latencies))
                        p95 = sorted_lats[int(len(sorted_lats)*0.95)]
                    else:
                        p95 = 0
                        
                    print(f"Progress: {total_processed_session}/{len(pending_rows)} | "
                          f"Concurrency: {current_concurrency} | "
                          f"Rate: {rate:.1f} req/s | "
                          f"Avg Lat: {avg_lat:.2f}s | p95 Lat: {p95:.2f}s | "
                          f"ETA: {eta:.0f}s")
                          
                if len(results_queue) >= 20:
                    for res in results_queue:
                        completed[str(res["row_idx"])] = res
                    with open(checkpoint_file, 'w') as f:
                        json.dump(completed, f)
                    results_queue.clear()
                    
            for res in results_queue:
                completed[str(res["row_idx"])] = res
            with open(checkpoint_file, 'w') as f:
                json.dump(completed, f)
                
        print("\nAll tasks completed. Writing to XLSM safely (single threaded)...")
        wb_out = openpyxl.load_workbook(self.input_file, keep_vba=True)
        ws_out = wb_out["Data"]
        max_col = ws_out.max_column
        
        col_map = {
            "Raw_text": 3,
            "Rating": 4,
            "Review_date": 5,
            "Reply_date": 6,
            "Support_reply": 7
        }
        
        audit_headers = [
            "Verification_Status", "Field_Affected", "Original_Value",
            "Trustpilot_Value", "Correction_Made", "Evidence_Source",
            "Confidence", "Reason_Action_Taken"
        ]
        
        audit_start = max_col
        if ws_out.cell(row=2, column=max_col).value != audit_headers[-1]:
            for i, h in enumerate(audit_headers, 1):
                ws_out.cell(row=2, column=max_col + i, value=h)
            audit_start = max_col
        else:
            audit_start = max_col - len(audit_headers)
            
        tally = {
            "Total_Rows": len(target_rows),
            "VERIFIED_MATCH": 0,
            "VERIFIED_AND_CORRECTED": 0,
            "MULTIPLE_POSSIBLE_MATCHES": 0,
            "REVIEW_NOT_FOUND": 0,
            "ACCESS_LIMITED": 0,
            "SOURCE_DATA_INSUFFICIENT": 0,
            "RATE_LIMITED": 0,
            "Corrections": { "Raw_text": 0, "Rating": 0, "Review_date": 0, "Reply_date": 0, "Support_reply": 0 },
            "Total_Retries_Errors": 0
        }
        
        for row_dict in target_rows:
            r_idx = str(row_dict["_row_idx"])
            if r_idx in completed:
                res = completed[r_idx]
                status_code = res["status_code"]
                audit_records = res["audit_records"]
                
                if status_code not in tally:
                    tally[status_code] = 0
                tally[status_code] += 1
                
                if res.get("http_status") not in ["SUCCESS", "REVIEW_NOT_FOUND"]:
                    tally["Total_Retries_Errors"] += 1
                
                for field, tp_val in res.get("corrections", {}).items():
                    if field in col_map:
                        ws_out.cell(row=int(r_idx), column=col_map[field], value=tp_val)
                        tally["Corrections"][field] += 1
                        
                row_obj_idx = int(r_idx)
                ws_out.cell(row=row_obj_idx, column=audit_start + 1, value=status_code)
                ws_out.cell(row=row_obj_idx, column=audit_start + 2, value=" | ".join(r.get("Field_Affected","") for r in audit_records))
                ws_out.cell(row=row_obj_idx, column=audit_start + 3, value=" | ".join(r.get("Original_Value","") for r in audit_records))
                ws_out.cell(row=row_obj_idx, column=audit_start + 4, value=" | ".join(r.get("Trustpilot_Value","") for r in audit_records))
                ws_out.cell(row=row_obj_idx, column=audit_start + 5, value=" | ".join(r.get("Correction_Made","") for r in audit_records))
                ws_out.cell(row=row_obj_idx, column=audit_start + 6, value=" | ".join(r.get("Evidence_Source","") for r in audit_records))
                ws_out.cell(row=row_obj_idx, column=audit_start + 7, value=" | ".join(r.get("Confidence","") for r in audit_records))
                ws_out.cell(row=row_obj_idx, column=audit_start + 8, value=" | ".join(r.get("Reason_Action_Taken","") for r in audit_records))
                
        print(f"Saving to {self.output_file}...")
        wb_out.save(self.output_file)
        
        return tally

    def verify_all(self):
        print(f"Loading {self.input_file} for data reading (data_only=True)...")
        # For reading, we use data_only=True to get actual cell values, not formulas
        wb_data = openpyxl.load_workbook(self.input_file, data_only=True)
        ws_data = wb_data["Data"]
        
        headers = [cell.value for cell in ws_data[2]]
        
        # Target rows finding
        target_rows = []
        for idx, row in enumerate(ws_data.iter_rows(min_row=3, values_only=True), start=3):
            row_dict = dict(zip(headers, row))
            if row_dict.get("Platform") == "Trustpilot" and row_dict.get("Url"):
                row_dict["_row_idx"] = idx
                target_rows.append(row_dict)

        print(f"Loading {self.input_file} for writing (keep_vba=True, data_only=False)...")
        # For writing, we load again with keep_vba=True to preserve structure and formulas
        wb_out = openpyxl.load_workbook(self.input_file, keep_vba=True)
        ws_out = wb_out["Data"]
        
        # Setup Audit Columns
        audit_headers = [
            "Verification_Status",
            "Field_Affected",
            "Original_Value",
            "Trustpilot_Value",
            "Correction_Made",
            "Evidence_Source",
            "Confidence",
            "Reason_Action_Taken"
        ]
        
        max_col = ws_out.max_column
        for i, h in enumerate(audit_headers, 1):
            ws_out.cell(row=2, column=max_col + i, value=h)
            
        results = []
        tally = {
            "Total_Rows": len(target_rows),
            "VERIFIED_MATCH": 0,
            "VERIFIED_AND_CORRECTED": 0,
            "MULTIPLE_POSSIBLE_MATCHES": 0,
            "REVIEW_NOT_FOUND": 0,
            "ACCESS_LIMITED": 0,
            "SOURCE_DATA_INSUFFICIENT": 0,
            "Corrections": {
                "Raw_text": 0,
                "Rating": 0,
                "Review_date": 0,
                "Reply_date": 0,
                "Support_reply": 0
            }
        }

        print(f"Processing {len(target_rows)} rows...")
        
        for i, r_data in enumerate(target_rows, 1):
            row_idx = r_data["_row_idx"]
            url = r_data["Url"]
            expected_id = self._extract_id_from_url(url)
            expected_text = r_data.get("Raw_text", "")
            
            print(f"[{i}/{len(target_rows)}] Row {row_idx} | ID: {expected_id}")
            
            success, status, content = self._fetch_jina_markdown(url)
            
            audit_records = []
            
            if success:
                is_match, reason = self._verify_identity(expected_id, expected_text, content)
                if is_match:
                    status_code = "VERIFIED_MATCH"
                    
                    # Extract and correct fields
                    tp_fields = self._safely_extract_fields(content)
                    
                    # Column mapping for original fields
                    col_map = {
                        "Raw_text": 3,
                        "Rating": 4,
                        "Review_date": 5,
                        "Reply_date": 6,
                        "Support_reply": 7
                    }
                    
                    row_corrected = False
                    for field, col_idx in col_map.items():
                        orig_val = r_data.get(field)
                        tp_val = tp_fields.get(field)
                        
                        if tp_val is not None:
                            orig_str = str(orig_val).strip().replace('\r\n', '\n') if orig_val else ""
                            tp_str = str(tp_val).strip()
                            
                            if orig_str != tp_str:
                                ws_out.cell(row=row_idx, column=col_idx, value=tp_val)
                                row_corrected = True
                                tally["Corrections"][field] += 1
                                audit_records.append({
                                    "Field_Affected": field,
                                    "Original_Value": str(orig_val)[:100],
                                    "Trustpilot_Value": str(tp_val)[:100],
                                    "Correction_Made": "Yes",
                                    "Evidence_Source": "Jina Markdown Extract",
                                    "Confidence": "High",
                                    "Reason_Action_Taken": f"Reliable Trustpilot value differed from Excel."
                                })
                                
                    if row_corrected:
                        status_code = "VERIFIED_AND_CORRECTED"
                    
                    if not audit_records:
                        audit_records.append({
                            "Field_Affected": "None",
                            "Original_Value": "N/A",
                            "Trustpilot_Value": "N/A",
                            "Correction_Made": "No",
                            "Evidence_Source": "Jina Markdown",
                            "Confidence": "High",
                            "Reason_Action_Taken": "Identity verified. No discrepancies found or missing fields not reliably extractable."
                        })
                else:
                    status_code = "MULTIPLE_POSSIBLE_MATCHES"
                    audit_records.append({
                        "Field_Affected": "None",
                        "Original_Value": "N/A",
                        "Trustpilot_Value": "N/A",
                        "Correction_Made": "No",
                        "Evidence_Source": "Jina Markdown",
                        "Confidence": "Low",
                        "Reason_Action_Taken": f"Content retrieved but identity unconfirmed. Reason: {reason}"
                    })
            elif status in ["ACCESS_LIMITED", "REVIEW_NOT_FOUND"]:
                status_code = status
                audit_records.append({
                    "Field_Affected": "None",
                    "Original_Value": "N/A",
                    "Trustpilot_Value": "N/A",
                    "Correction_Made": "No",
                    "Evidence_Source": "HTTP Response",
                    "Confidence": "High",
                    "Reason_Action_Taken": f"Failed to retrieve content. Status: {status}"
                })
            
            tally[status_code] += 1
            
            # Write audit records (write first record on the row, subsequent records on new lines in the cell or just pick the first)
            # Since multiple fields can be corrected, we format the audit columns as a joined string of all records for the row
            ws_out.cell(row=row_idx, column=max_col + 1, value=status_code)
            ws_out.cell(row=row_idx, column=max_col + 2, value=" | ".join(r["Field_Affected"] for r in audit_records))
            ws_out.cell(row=row_idx, column=max_col + 3, value=" | ".join(r["Original_Value"] for r in audit_records))
            ws_out.cell(row=row_idx, column=max_col + 4, value=" | ".join(r["Trustpilot_Value"] for r in audit_records))
            ws_out.cell(row=row_idx, column=max_col + 5, value=" | ".join(r["Correction_Made"] for r in audit_records))
            ws_out.cell(row=row_idx, column=max_col + 6, value=" | ".join(r["Evidence_Source"] for r in audit_records))
            ws_out.cell(row=row_idx, column=max_col + 7, value=" | ".join(r["Confidence"] for r in audit_records))
            ws_out.cell(row=row_idx, column=max_col + 8, value=" | ".join(r["Reason_Action_Taken"] for r in audit_records))
            
            time.sleep(1) # Courtesy delay
            
        print(f"Saving to {self.output_file}...")
        wb_out.save(self.output_file)
        
        return tally
